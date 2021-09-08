# -*- coding: utf-8 -*-
"""
Created on Sat Jul  3 19:26:40 2021

@author: Thales
"""

import pandas as pd
from datetime import timedelta
from datetime import datetime
from datetime import date
from timeit import default_timer as timer
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def realmenteSuspeitos(tabelaAssessor):
    """
    Parameters
    ----------
    tabelaAssessor : DataFrame
        Tabela Assessor

    Returns
    -------
    tabelaTotal : DataFrame
        Tabela Assessor, porém com apenas os "Realmente Suspeitos", de acordo com a regra de negócio

    """
    print("Iniciando 'Realmente Suspeitos'...")
    paramColunaIdAssessor = 2
    paramColunaDataNotifAssessor = 11
    tabelaTotal = tabelaAssessor
    tabelaSuspeitos = tabelaTotal.where((tabelaTotal["Data da Notificação"] >= datetime.strptime('01-06-2020', '%d-%m-%Y')) & (tabelaTotal["Situação"] == "SUSPEITA")).dropna(how='all').sort_values(by=["Cód. Paciente", "Data da Notificação"]).drop_duplicates("Cód. Paciente", ignore_index=True)
    startProcessamento = timer()
    porcAtual = 0
    somaVeloc = 0
    countLinhas = 0
    qtdNotif = len(tabelaSuspeitos)
    porcentil = round(qtdNotif / 100)
    for row in tabelaSuspeitos.itertuples(): #Passando por cada suspeito
        situacoes = tabelaTotal.where(tabelaTotal["Cód. Paciente"] == row[paramColunaIdAssessor]).dropna(how='all') #Pega apenas a situação e data de notificação de cada registro que o suspeito possui na lista total
        flag = False
        #print("Teste de situação: " + situacoes["Situação"].values)
        if "CONFIRMADO" in situacoes["Situação"].values: #Se houver um confirmado na lista de situações
            confirmadosAssessor = situacoes.where(situacoes["Situação"] == "CONFIRMADO").dropna(how='all')
            for rowConfirmado in confirmadosAssessor.itertuples(): #Passa por cada situação
                flag = False
                if rowConfirmado[paramColunaDataNotifAssessor] >= (row[paramColunaDataNotifAssessor] - timedelta(days=90)):
                    tabelaSuspeitos.drop(row.Index, inplace=True) #Tira o suspeito da lista de suspeitos
                    flag = True
                    break
            if not flag:
                if "NEGATIVO" in situacoes["Situação"].values: #Se houver um negativo nas situações
                    negativosAssessor = situacoes.where(situacoes["Situação"] == "NEGATIVO").dropna(how='all')
                    for rowNegativo in negativosAssessor.itertuples():                
                        if rowNegativo[paramColunaDataNotifAssessor] >= (row[paramColunaDataNotifAssessor] - timedelta(days=2)):                    
                            tabelaSuspeitos.drop(row.Index, inplace=True) #Tira o suspeito da lista de suspeitos
                            break
                        
        countLinhas += 1
        if countLinhas >= (porcAtual + 1) * porcentil:
            porcAtual += 1
            os.system('cls')
            porcTimer = timer()
            tempoAtual = porcTimer - startProcessamento
            velocEstimada = countLinhas / tempoAtual
            somaVeloc += velocEstimada
            velocMedia = somaVeloc / porcAtual
            tempoEstimadoTotal = (qtdNotif - countLinhas) / velocEstimada
            stringTempoEstimado = segundoEmHora(tempoEstimadoTotal)
            print("REALMENTE SUSPEITOS\nConcluido: " + str(porcAtual) + "% aos " + "{:.0f}".format(tempoAtual) + " segundos (desde o início do laço).\n" + str(countLinhas) + " linhas lidas de " + str(qtdNotif) + " total.\nVelocidade estimada: " + "{:.2f}".format(velocEstimada) + " linhas por segundo.\nVelocidade media: " + "{:.2f}".format(velocMedia) + " linhas por segundo.\nTempo estimado: " + stringTempoEstimado + " para conlusão\n\n")
    
    tabelaTotal = tabelaTotal.where(tabelaTotal["Situação"] != "SUSPEITA").dropna(how='all')
    tabelaTotal = pd.concat([tabelaTotal, tabelaSuspeitos])
    # with pd.ExcelWriter('Total e Realmente Suspeitos ' + paramDataAtual.strftime("%d.%m") + '.xlsx') as writer:
    #     tabelaTotal.to_excel(writer, index=False)
    print("Total de " + str(len(tabelaTotal)) + " total e realmente suspeitos.")
    return tabelaTotal

def segundoEmHora(segundos):
    """
    Parameters
    ----------
    segundos : Float
        Quantidade de segundos a serem convertidos em String de Tempo

    Returns
    -------
    stringTempo : String
        String representando os segundos recebidos em Hora:Minuto:Segundo

    """
    horas = int(segundos / 3600) if int(segundos / 3600) >= 10 else "0" + str(int(segundos / 3600))
    minutos = int((segundos % 3600) / 60) if int((segundos % 3600) / 60) >= 10 else "0" + str(int((segundos % 3600) / 60))
    segundos = int((segundos % 3600) % 60) if int((segundos % 3600) % 60) >= 10 else "0" + str(int((segundos % 3600) % 60))
    stringTempo = str(horas) + ":" + str(minutos) + ":" + str(segundos)
    return stringTempo

def bordaCelula(border_style, color):
    borda = Border(left=Side(border_style=border_style, color=color),
                             right=Side(border_style=border_style, color=color),
                             top=Side(border_style=border_style, color=color),
                             bottom=Side(border_style=border_style, color=color))
    return borda

def autoSizeColumn(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[1].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length * 1.2
        worksheet.column_dimensions[column].width = adjusted_width
    return worksheet

def consolidadoTestes(listaUnidades):
    wb = Workbook()
    ws = wb.active
    
    offsetLinha = 2
    
    ws['B1'] = "Essa semana (" + paramDataAtras.strftime("%d/%m/%Y") + " a " + paramDataAtual.strftime("%d/%m/%Y") + ")"
    ws['B1'].font = Font(bold=True)
    ws['B1'].alignment = Alignment(horizontal='center')  
    ws.merge_cells('B1:D1')
    ws['B1'].border = bordaCelula('thick', '00000000')
    ws['C1'].border = bordaCelula('thick', '00000000')
    ws['D1'].border = bordaCelula('thick', '00000000')
    
    ws['E1'] = "Mês Passado (" + paramPrimeiroDiaMesPassado.strftime("%d/%m/%Y") + " a " + paramUltimoDiaMesPassado.strftime("%d/%m/%Y") + ")"
    ws['E1'].font = Font(bold=True)
    ws['E1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('E1:G1')
    ws['E1'].border = bordaCelula('thick', '00000000')
    ws['F1'].border = bordaCelula('thick', '00000000')
    ws['G1'].border = bordaCelula('thick', '00000000')
    
    ws['H1'] = "TOTAL (01/01/2020 a " + paramDataAtual.strftime("%d/%m/%Y") + ")"
    ws['H1'].font = Font(bold=True)
    ws['H1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('H1:J1')
    ws['H1'].border = bordaCelula('thick', '00000000')
    ws['I1'].border = bordaCelula('thick', '00000000')
    ws['J1'].border = bordaCelula('thick', '00000000')    
    
    
    ws['A2'] = "Unidade"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].border = bordaCelula('thick', '00000000')
    ws['A2'].font = Font(bold=True)
    ws['B2'] = "Testes Realizados"
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['B2'].border = bordaCelula('thick', '00000000')
    ws['B2'].font = Font(bold=True)
    ws['C2'] = "Dispensações sem Resultado"
    ws['C2'].alignment = Alignment(horizontal='center')
    ws['C2'].border = bordaCelula('thick', '00000000')
    ws['C2'].font = Font(bold=True)
    ws['D2'] = "% dos testes realizados"
    ws['D2'].alignment = Alignment(horizontal='center')
    ws['D2'].border = bordaCelula('thick', '00000000')
    ws['D2'].font = Font(bold=True)
    ws['E2'] = "Testes Realizados"
    ws['E2'].alignment = Alignment(horizontal='center')
    ws['E2'].border = bordaCelula('thick', '00000000')
    ws['E2'].font = Font(bold=True)
    ws['F2'] = "Dispensações sem Resultado"
    ws['F2'].alignment = Alignment(horizontal='center')
    ws['F2'].border = bordaCelula('thick', '00000000')
    ws['F2'].font = Font(bold=True)
    ws['G2'] = "% dos testes realizados"
    ws['G2'].alignment = Alignment(horizontal='center')
    ws['G2'].border = bordaCelula('thick', '00000000')
    ws['G2'].font = Font(bold=True)
    ws['H2'] = "Testes Realizados"
    ws['H2'].alignment = Alignment(horizontal='center')
    ws['H2'].border = bordaCelula('thick', '00000000')
    ws['H2'].font = Font(bold=True)
    ws['I2'] = "Dispensações sem Resultado"
    ws['I2'].alignment = Alignment(horizontal='center')
    ws['I2'].border = bordaCelula('thick', '00000000')
    ws['I2'].font = Font(bold=True)
    ws['J2'] = "% dos testes realizados"
    ws['J2'].alignment = Alignment(horizontal='center')
    ws['J2'].border = bordaCelula('thick', '00000000')
    ws['J2'].font = Font(bold=True)
    
    contLinha = offsetLinha + 1
    
    for unidade in listaUnidades:
        if unidade['Apelido'] == "Postão":
            continue
        stringLinha = str(contLinha)
        
        #Dados do período semana
        ws['A' + stringLinha] = unidade['Apelido']
        ws['A' + stringLinha].border = Border(right=Side(style='thick', color='00000000'),
                                              left=Side(style='thick', color='00000000'),
                                              top=Side(style='thin', color='00000000'),
                                              bottom=Side(style='thin', color='00000000'),)
        
        ws['B' + stringLinha] = unidade['Testes Dispensados Semana']
        ws['B' + stringLinha].border = bordaCelula('thin', '00000000')
        
        ws['C' + stringLinha] = unidade['Dispensações sem Resultado Semana']
        ws['C' + stringLinha].border = bordaCelula('thin', '00000000')
        
        if ws['B' + stringLinha].value > 0:
            ws['D' + stringLinha] = "{:.2f}".format((ws['C' + stringLinha].value / ws['B' + stringLinha].value) * 100) + "%"
        else:
            ws['D' + stringLinha] = '-'        
        ws['D' + stringLinha].border = bordaCelula('thin', '00000000')  
        if ws['D' + stringLinha].value != '-' and (ws['C' + stringLinha].value / ws['B' + stringLinha].value) * 100 > paramPorcTestes: 
            ws['D' + stringLinha].font = Font(color="00FF0000", bold=True)
        
        #Dados do período mês passado
        ws['E' + stringLinha] = unidade['Testes Dispensados Mes Passado']
        ws['E' + stringLinha].border = Border(left=Side(style='thick', color='00000000'),
                                              right=Side(style='thin', color='00000000'),
                                              top=Side(style='thin', color='00000000'),
                                              bottom=Side(style='thin', color='00000000'),)
        
        ws['F' + stringLinha] = unidade['Dispensações sem Resultado Mes Passado']
        ws['F' + stringLinha].border = bordaCelula('thin', '00000000')
        
        if ws['E' + stringLinha].value > 0:
            ws['G' + stringLinha] = "{:.2f}".format((ws['F' + stringLinha].value / ws['E' + stringLinha].value) * 100) + "%"
        else:
            ws['G' + stringLinha] = '-'
        if ws['G' + stringLinha].value != '-' and (ws['F' + stringLinha].value / ws['E' + stringLinha].value) * 100 > paramPorcTestes:
            ws['G' + stringLinha].font = Font(color="00FF0000", bold=True)
        ws['G' + stringLinha].border = Border(right=Side(style='thick', color='00000000'),
                                              left=Side(style='thin', color='00000000'),
                                              top=Side(style='thin', color='00000000'),
                                              bottom=Side(style='thin', color='00000000'),) 
        
        #Dados do período total
        ws['H' + stringLinha] = unidade['Testes Dispensados Total']
        ws['H' + stringLinha].border = Border(left=Side(style='thick', color='00000000'),
                                              right=Side(style='thin', color='00000000'),
                                              top=Side(style='thin', color='00000000'),
                                              bottom=Side(style='thin', color='00000000'),)
        
        ws['I' + stringLinha] = unidade['Dispensações sem Resultado Total']
        ws['I' + stringLinha].border = bordaCelula('thin', '00000000')
        
        if ws['H' + stringLinha].value > 0:
            ws['J' + stringLinha] = "{:.2f}".format((ws['I' + stringLinha].value / ws['H' + stringLinha].value) * 100) + "%"
        else:
            ws['J' + stringLinha] = '-'
        if ws['J' + stringLinha].value != '-' and (ws['I' + stringLinha].value / ws['H' + stringLinha].value) * 100 > paramPorcTestes:
            ws['J' + stringLinha].font = Font(color="00FF0000", bold=True)
        ws['J' + stringLinha].border = Border(right=Side(style='thick', color='00000000'),
                                              left=Side(style='thin', color='00000000'),
                                              top=Side(style='thin', color='00000000'),
                                              bottom=Side(style='thin', color='00000000'),) 
        
        ws['A' + stringLinha].alignment = Alignment(horizontal='center')
        ws['B' + stringLinha].alignment = Alignment(horizontal='center')
        ws['C' + stringLinha].alignment = Alignment(horizontal='center')
        ws['D' + stringLinha].alignment = Alignment(horizontal='center')
        ws['E' + stringLinha].alignment = Alignment(horizontal='center')
        ws['F' + stringLinha].alignment = Alignment(horizontal='center')
        ws['G' + stringLinha].alignment = Alignment(horizontal='center')
        ws['H' + stringLinha].alignment = Alignment(horizontal='center')
        ws['I' + stringLinha].alignment = Alignment(horizontal='center')
        ws['J' + stringLinha].alignment = Alignment(horizontal='center')
        
        contLinha += 1
        
    ws['A' + str(contLinha)].border = Border(top=Side(style='thick', color="00000000"))
    ws['B' + str(contLinha)].border = Border(top=Side(style='thick', color="00000000"))
    ws['C' + str(contLinha)].border = Border(top=Side(style='thick', color="00000000"))
    ws['D' + str(contLinha)].border = Border(top=Side(style='thick', color="00000000"))
    ws['E' + str(contLinha)].border = Border(top=Side(style='thick', color="00000000"))
    ws['F' + str(contLinha)].border = Border(top=Side(style='thick', color="00000000"))
    ws['G' + str(contLinha)].border = Border(top=Side(style='thick', color="00000000"))
    ws['H' + str(contLinha)].border = Border(top=Side(style='thick', color="00000000"))
    ws['I' + str(contLinha)].border = Border(top=Side(style='thick', color="00000000"))
    ws['J' + str(contLinha)].border = Border(top=Side(style='thick', color="00000000"))
    
    ws = autoSizeColumn(ws)
    wb.save('Unidades/Consolidado Testes Rápidos ' + paramDataAtras.strftime("%d.%m") + " a " + paramDataAtual.strftime("%d.%m") + '.xlsx')

listaUnidades = [
    {'CNES':2064081, 'Nome':'UBS DR MILTON BARONI DE BARRETOS', 'Apelido': 'América'},
    {'CNES':2053314, 'Nome':'UNIDADE BASICA DE SAUDE DR LOTFALLAH MIZIARA', 'Apelido': 'Barretos 2'},
    {'CNES':2093642, 'Nome':'UNIDADE BASICA DE SAUDE DR WILSON HAYEK SAIHG', 'Apelido': 'Cecapinha'},
    {'CNES':2061473, 'Nome':'UNIDADE BASICA DE SAUDE DR PAULO PRATA', 'Apelido': 'Cristiano de Carvalho'},
    {'CNES':2064103, 'Nome':'UBS DR ALLY ALAHMAR DE BARRETOS', 'Apelido': 'CSU'},
    {'CNES':2093650, 'Nome':'UNIDADE BASICA DE SAUDE DR BARTOLOMEU MARAGLIANO V', 'Apelido': 'Derby'},
    {'CNES':2048744, 'Nome':'UNIDADE BASICA DR JOSE PARASSU CARVALHO', 'Apelido': 'Ibirapuera'},
    {'CNES':2035731, 'Nome':'UNIDADE BASICA DE SAUDE DO IBITU', 'Apelido': 'Ibitu'},
    {'CNES':2048736, 'Nome':'UNIDADE BASICA DE SAUDE DR APOLONIO MORAES E SOUZA', 'Apelido': 'Los Angeles'},
    {'CNES':2784580, 'Nome':'UNIDADE BASICA DE SAUDE DR SERGIO PIMENTA', 'Apelido': 'Marilia'},
    {'CNES':7585020, 'Nome':'USF DO BAIRRO NOVA BARRETOS', 'Apelido': 'Nova Barretos'},
    {'CNES':2056062, 'Nome':'UNIDADE BASICA ARCHIMEDES MACHADO DE BARRETOS', 'Apelido': 'Pimenta'},
    {'CNES':2784572, 'Nome':'AMBULATORIO DE ESPECIALIDADES ARE I', 'Apelido': 'Postão'},
    {'CNES':2784599, 'Nome':'UNIDADE BASICA DE SAUDE FRANCOLINO GALVAO DE SOUZA', 'Apelido': 'São Francisco'},
    {'CNES':7565577, 'Nome':'AMBULATORIO SAUDE DO IDOSO', 'Apelido': 'Saúde do Idoso'},
    {'CNES':7122217, 'Nome':'UNIDADE ESF DR LUIZ SPINA', 'Apelido': 'Spina'},
    {'CNES':7035861, 'Nome':'UPA 24H ZAID ABRAO GERAIGE', 'Apelido': 'UPA'}
]

codigosVacina = [17969, 17970, 17971, 17972]

paramDataAtual = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
paramDataTruncAtual = date.today()
paramTamanhoSemana = 7
paramDataAtras = paramDataAtual - timedelta(days=7)
paramUltimoDiaMesPassado = paramDataAtual - timedelta(days=paramDataAtual.day)
paramPrimeiroDiaMesPassado = paramUltimoDiaMesPassado - timedelta(days=paramUltimoDiaMesPassado.day - 1)
paramPorcTestes = 10 #Parâmetro que considera a partir de que % é considerado um alerta (ex: Acima de 10% será pintado de vermelho)
filtroTeste1 = None
filtroTeste2 = None
filtroCodsTeste = None

paramCodigoSaidaVacina = 15

def qtdAgravos(situacao, unidade, semana, mes): #Retorna a quantidade de suspeitas da unidade, total ou da semana
    filtro = (listaTotalAssessor["Situação"] == situacao) & (listaTotalAssessor["Unidade"] == unidade["Nome"])
    if semana:
        filtro = filtro & (listaTotalAssessor["Data da Digitação"] <= paramDataAtual) & (listaTotalAssessor["Data da Digitação"] >= paramDataAtual - timedelta(days=paramTamanhoSemana))
    elif mes:
        filtro = filtro & (listaTotalAssessor["Data da Digitação"] <= paramUltimoDiaMesPassado) & (listaTotalAssessor["Data da Digitação"] >= paramPrimeiroDiaMesPassado)
    agravosUnidade = listaTotalAssessor.where(filtro).dropna(how='all')
    return len(agravosUnidade)

def semInicioSintomas(unidade, semana, mes): #Retorna a quantidade de notificações (suspeitas e confirmadas) sem data de inicio de sintomas da unidade, total ou da semana
    filtro = ((listaTotalAssessor["Situação"] == "SUSPEITA") | (listaTotalAssessor["Situação"] == "CONFIRMADO")) & (listaTotalAssessor["Unidade"] == unidade["Nome"]) & (pd.isnull(listaTotalAssessor["Data dos Primeiros Sintomas"]))
    if semana:
        filtro = filtro & (listaTotalAssessor["Data da Digitação"] <= paramDataAtual) & (listaTotalAssessor["Data da Digitação"] >= paramDataAtual - timedelta(days=paramTamanhoSemana))
    elif mes:
        filtro = filtro & (listaTotalAssessor["Data da Digitação"] <= paramUltimoDiaMesPassado) & (listaTotalAssessor["Data da Digitação"] >= paramPrimeiroDiaMesPassado)
    semSintomasUnidade = listaTotalAssessor.where(filtro).dropna(how='all')
    return len(semSintomasUnidade)

def difDigitNotif(unidade, semana, mes): #Retorna a quantidade de notificações com diferença entre Data de Digitação e Data de Notificação, total ou da semana
    filtro = (listaTotalAssessor["Situação"] != "DESCARTADO") & (listaTotalAssessor["Unidade"] == unidade["Nome"]) & (listaTotalAssessor["Data da Digitação"] != listaTotalAssessor["Data da Notificação"])
    if semana:
        filtro = filtro & (listaTotalAssessor["Data da Digitação"] <= paramDataAtual) & (listaTotalAssessor["Data da Digitação"] >= paramDataAtual - timedelta(days=paramTamanhoSemana))
    elif mes:
        filtro = filtro & (listaTotalAssessor["Data da Digitação"] <= paramUltimoDiaMesPassado) & (listaTotalAssessor["Data da Digitação"] >= paramPrimeiroDiaMesPassado)
    diferenca = listaTotalAssessor.where(filtro).dropna(how='all')
    return len(diferenca)

def dispSemAgravo(unidade, semana, mes):
    filtro = listaDispSemAgravo["COD. UNID."] == unidade["CNES"]
    if semana:
        filtro = filtro & (listaDispSemAgravo["DATA DISP."] <= paramDataAtual) & (listaDispSemAgravo["DATA DISP."] >= paramDataAtual - timedelta(days=paramTamanhoSemana))
    elif mes:
        filtro = filtro & (listaDispSemAgravo["DATA DISP."] <= paramUltimoDiaMesPassado) & (listaDispSemAgravo["DATA DISP."] >= paramPrimeiroDiaMesPassado)
    semAgravo = listaDispSemAgravo.where(filtro).dropna(how='all')
    return len(semAgravo)

def dispensacoes(unidade, semana, mes):
    filtro = listaTotalDispensacoes["COD. UNID."] == unidade["CNES"]
    if semana:
        filtro = filtro & (listaTotalDispensacoes["DATA DISP."] <= paramDataAtual) & (listaTotalDispensacoes["DATA DISP."] >= paramDataAtual - timedelta(days=paramTamanhoSemana))
    elif mes:
        filtro = filtro & (listaTotalDispensacoes["DATA DISP."] <= paramUltimoDiaMesPassado) & (listaTotalDispensacoes["DATA DISP."] >= paramPrimeiroDiaMesPassado)
    totalDispensacoes = listaTotalDispensacoes.where(filtro).dropna(how='all')
    return totalDispensacoes["QTD DISP."].sum()

def vacivida(unidade, semana, mes):
    filtro = listaVacivida["CNES_ESTABELECIMENTO"] == unidade["CNES"]
    if semana:
        filtro = filtro & (listaVacivida["DATA_APLICACAO_VACINA"] <= paramDataAtual) & (listaVacivida["DATA_APLICACAO_VACINA"] >= paramDataAtual - timedelta(days=paramTamanhoSemana))
    elif mes:
        filtro = filtro & (listaVacivida["DATA_APLICACAO_VACINA"] <= paramUltimoDiaMesPassado) & (listaVacivida["DATA_APLICACAO_VACINA"] >= paramPrimeiroDiaMesPassado)
    retornoVacivida = listaVacivida.where(filtro).dropna(how='all')
    return len(retornoVacivida)

def saidaVacina(unidade, tipoSaida, semana, mes):
    filtro = (listaTotalSaidas["CÓD UNID."] == unidade["CNES"]) & (listaTotalSaidas["TIPO MOVIMENTAÇÃO"] == "SAÍDAS DIVERSAS") & (listaTotalSaidas["TIPO"] == tipoSaida)
    filtroCods = None
    for cod in codigosVacina:
        if filtroCods is None:
            filtroCods = (listaTotalSaidas["CÓD MAT."] == cod)
        else:
            filtroCods = filtroCods | (listaTotalSaidas["CÓD MAT."] == cod)
    filtro = filtro & filtroCods
    if semana:
        filtro = filtro & (listaTotalSaidas["DATA MOV"] <= paramDataAtual) & (listaTotalSaidas["DATA MOV"] >= paramDataAtual - timedelta(days=paramTamanhoSemana))    
    elif mes:
        filtro = filtro & (listaTotalSaidas["DATA MOV"] <= paramUltimoDiaMesPassado) & (listaTotalSaidas["DATA MOV"] >= paramPrimeiroDiaMesPassado)    
    saidas = listaTotalSaidas.where(filtro).dropna(how='all')
    return saidas["QTD"].sum()
    
listaTotalAssessor = pd.read_excel("lista total assessor.xlsx")
listaTotalAssessor = realmenteSuspeitos(listaTotalAssessor)

listaDispSemAgravo = pd.read_excel("disp sem agravo.xlsx").drop_duplicates("ID DISP.")
listaDispSemAgravo["DATA DISP."] = pd.to_datetime(listaDispSemAgravo["DATA DISP."], dayfirst=True)

listaTotalDispensacoes = pd.read_excel("lista total dispensacoes.xlsx")
listaTotalDispensacoes = listaTotalDispensacoes.where(listaTotalDispensacoes["DATA DISP."] != "DATA DISP.").dropna(how='all')
listaTotalDispensacoes["DATA DISP."] = pd.to_datetime(listaTotalDispensacoes["DATA DISP."], dayfirst=True)

listaVacivida = pd.read_excel("vacivida.xlsx")
listaVacivida = listaVacivida.where(listaVacivida["DATA_APLICACAO_VACINA"] != "DATA_APLICACAO_VACINA")
listaVacivida["DATA_APLICACAO_VACINA"] = pd.to_datetime(listaVacivida["DATA_APLICACAO_VACINA"], dayfirst=True)

#print("Comecei a ler a lista de saidas")
listaTotalSaidas = pd.read_excel("lista total saidas.xlsx")
listaTotalSaidas["DATA MOV"] = pd.to_datetime(listaTotalSaidas["DATA MOV"], dayfirst=True)
print("Iniciando processamento...")

#print("Dia de hoje: " + str(paramDataAtual) + "\n" + str(paramTamanhoSemana) + " dias atrás: " + str(paramDataAtual - timedelta(days=paramTamanhoSemana)))

for unidade in listaUnidades:
    unidade["Suspeitas Total"] = qtdAgravos("SUSPEITA", unidade, False, False)
    unidade["Suspeitas Semana"] = qtdAgravos("SUSPEITA", unidade, True, False)
    unidade["Suspeitas Mes Passado"] = qtdAgravos("SUSPEITA", unidade, False, True)
    unidade["Positivos Total"] = qtdAgravos("CONFIRMADO", unidade, False, False)
    unidade["Positivos Semana"] = qtdAgravos("CONFIRMADO", unidade, True, False)
    unidade["Positivos Mes Passado"] = qtdAgravos("CONFIRMADO", unidade, False, True)
    unidade["Negativos Total"] = qtdAgravos("NEGATIVO", unidade, False, False)
    unidade["Negativos Semana"] = qtdAgravos("NEGATIVO", unidade, True, False)
    unidade["Negativos Mes Passado"] = qtdAgravos("NEGATIVO", unidade, False, True)
    unidade["Descartados Total"] = qtdAgravos("DESCARTADO", unidade, False, False)
    unidade["Descartados Semana"] = qtdAgravos("DESCARTADO", unidade, True, False)
    unidade["Descartados Mes Passado"] = qtdAgravos("DESCARTADO", unidade, False, True)
    unidade["Sem Sintomas Total"] = semInicioSintomas(unidade, False, False)
    unidade["Sem Sintomas Semana"] = semInicioSintomas(unidade, True, False)
    unidade["Sem Sintomas Mes Passado"] = semInicioSintomas(unidade, False, True)
    unidade["Diferença Digitação x Notificação Total"] = difDigitNotif(unidade, False, False)
    unidade["Diferença Digitação x Notificação Semana"] = difDigitNotif(unidade, True, False)
    unidade["Diferença Digitação x Notificação Mes Passado"] = difDigitNotif(unidade, False, True)
    unidade["Dispensações sem Resultado Total"] = dispSemAgravo(unidade, False, False)
    unidade["Dispensações sem Resultado Semana"] = dispSemAgravo(unidade, True, False)
    unidade["Dispensações sem Resultado Mes Passado"] = dispSemAgravo(unidade, False, True)
    unidade["Testes Dispensados Total"] = dispensacoes(unidade, False, False)
    unidade["Testes Dispensados Semana"] = dispensacoes(unidade, True, False)
    unidade["Testes Dispensados Mes Passado"] = dispensacoes(unidade, False, True)
    unidade["Vacivida Total"] = vacivida(unidade, False, False)
    unidade["Vacivida Semana"] = vacivida(unidade, True, False)
    unidade["Vacivida Mes Passado"] = vacivida(unidade, False, True)
    unidade["Saída de Vacina (Aplicação) Total"] = saidaVacina(unidade, paramCodigoSaidaVacina, False, False)
    unidade["Saída de Vacina (Aplicação) Semana"] = saidaVacina(unidade, paramCodigoSaidaVacina, True, False)
    unidade["Saída de Vacina (Aplicação) Mes Passado"] = saidaVacina(unidade, paramCodigoSaidaVacina, False, True)
    unidade["Aplicação Vacina Assessor - Vacivida (Semana)"] = unidade["Saída de Vacina (Aplicação) Semana"] -  unidade["Vacivida Semana"]
    unidade["Aplicação Vacina Assessor - Vacivida (Mes Passado)"] = unidade["Saída de Vacina (Aplicação) Mes Passado"] -  unidade["Vacivida Mes Passado"]
    
    wb = Workbook()
    ws = wb.active
    ws['A1'] = str(unidade["CNES"]) + " " + unidade["Nome"] + " - " + unidade["Apelido"]
    ws.merge_cells('A1:F1')
    ws['A3'] = "Dispensações sem Resultado (Total)"
    ws['B3'] = unidade["Dispensações sem Resultado Total"]
    ws['A4'] = "Dispensações sem Resultado (Mes Passado)"
    ws['B4'] = unidade["Dispensações sem Resultado Mes Passado"]
    if unidade["Testes Dispensados Mes Passado"] > 0:
        ws['C4'] = "{:.2f}".format((unidade["Dispensações sem Resultado Mes Passado"] / unidade["Testes Dispensados Mes Passado"]) * 100) + "%"
    ws['A5'] = "Dispensações sem Resultado (Semana)"
    ws['B5'] = unidade["Dispensações sem Resultado Semana"]
    if unidade["Testes Dispensados Semana"] > 0:
        ws['C5'] = "{:.2f}".format((unidade["Dispensações sem Resultado Semana"] / unidade["Testes Dispensados Semana"]) * 100) + "%"
    
    
    ws['A7'] = "Notificações sem Data de Inicio de Sintomas (Total)"
    ws['B7'] = unidade["Sem Sintomas Total"]
    ws['A8'] = "Notificações sem Data de Inicio de Sintomas (Mes Passado)"
    ws['B8'] = unidade["Sem Sintomas Mes Passado"]
    ws['A9'] = "Notificações sem Data de Inicio de Sintomas (Semana)"
    ws['B9'] = unidade["Sem Sintomas Semana"]
    
    ws['A11'] = "Notificações com diferença entre Data de Notificação e Data de Digitação (Total)"
    ws['B11'] = unidade["Diferença Digitação x Notificação Total"]
    ws['A12'] = "Notificações com diferença entre Data de Notificação e Data de Digitação (Mes Passado)"
    ws['B12'] = unidade["Diferença Digitação x Notificação Mes Passado"]
    ws['A13'] = "Notificações com diferença entre Data de Notificação e Data de Digitação (Semana)"
    ws['B13'] = unidade["Diferença Digitação x Notificação Semana"]
    
    ws['A15'] = "Notificações (Total)"
    ws['B15'] = "Suspeitas"
    ws['C15'] = "Positivas"
    ws['D15'] = "Negativas"
    ws['E15'] = "Descartadas"
    ws['F15'] = "Total"    
    ws['B16'] = unidade["Suspeitas Total"]
    ws['C16'] = unidade["Positivos Total"]
    ws['D16'] = unidade["Negativos Total"]
    ws['E16'] = unidade["Descartados Total"]
    somaNotifTotal = unidade["Suspeitas Total"] + unidade["Positivos Total"] + unidade["Negativos Total"] + unidade["Descartados Total"]    
    if somaNotifTotal > 0:
        porcSuspeitasTotal = (unidade["Suspeitas Total"] / somaNotifTotal) * 100
        porcPositivasTotal = (unidade["Positivos Total"] / somaNotifTotal) * 100
        porcNegativasTotal = (unidade["Negativos Total"] / somaNotifTotal) * 100
        porcDescartadasTotal = (unidade["Descartados Total"] / somaNotifTotal) * 100
        ws['B17'] = "{:.2f}".format(porcSuspeitasTotal) + "%"
        ws['C17'] = "{:.2f}".format(porcPositivasTotal) + "%"
        ws['D17'] = "{:.2f}".format(porcNegativasTotal) + "%"
        ws['E17'] = "{:.2f}".format(porcDescartadasTotal) + "%"
    ws['F16'] = somaNotifTotal
    
    ws['A19'] = "Notificações (Mes Passado)"
    ws['B19'] = "Suspeitas"
    ws['C19'] = "Positivas"
    ws['D19'] = "Negativas"
    ws['E19'] = "Descartadas"
    ws['F19'] = "Total"
    ws['B20'] = unidade["Suspeitas Mes Passado"]
    ws['C20'] = unidade["Positivos Mes Passado"]
    ws['D20'] = unidade["Negativos Mes Passado"]
    ws['E20'] = unidade["Descartados Mes Passado"]    
    somaNotifMesPassado = unidade["Suspeitas Mes Passado"] + unidade["Positivos Mes Passado"] + unidade["Negativos Mes Passado"] + unidade["Descartados Mes Passado"]
    if somaNotifMesPassado > 0:
        porcSuspeitasMesPassado = (unidade["Suspeitas Mes Passado"] / somaNotifMesPassado) * 100    
        porcPositivasMesPassado = (unidade["Positivos Mes Passado"] / somaNotifMesPassado) * 100
        porcNegativasMesPassado = (unidade["Negativos Mes Passado"] / somaNotifMesPassado) * 100
        porcDescartadasMesPassado = (unidade["Descartados Mes Passado"] / somaNotifMesPassado) * 100    
        ws['B21'] = "{:.2f}".format(porcSuspeitasMesPassado) + "%"
        ws['C21'] = "{:.2f}".format(porcPositivasMesPassado) + "%"
        ws['D21'] = "{:.2f}".format(porcNegativasMesPassado) + "%"    
        ws['E21'] = "{:.2f}".format(porcDescartadasMesPassado) + "%"
    ws['F20'] = somaNotifMesPassado
    
    ws['A23'] = "Notificações (Semana)"
    ws['B23'] = "Suspeitas"
    ws['C23'] = "Positivas"
    ws['D23'] = "Negativas"
    ws['E23'] = "Descartadas"
    ws['F23'] = "Total"
    ws['B24'] = unidade["Suspeitas Semana"]
    ws['C24'] = unidade["Positivos Semana"]
    ws['D24'] = unidade["Negativos Semana"]
    ws['E24'] = unidade["Descartados Semana"]    
    somaNotifSemana = unidade["Suspeitas Semana"] + unidade["Positivos Semana"] + unidade["Negativos Semana"] + unidade["Descartados Semana"]        
    if somaNotifSemana > 0:
        porcSuspeitasSemana = (unidade["Suspeitas Semana"] / somaNotifSemana) * 100    
        porcPositivasSemana = (unidade["Positivos Semana"] / somaNotifSemana) * 100
        porcNegativasSemana = (unidade["Negativos Semana"] / somaNotifSemana) * 100
        porcDescartadasSemana = (unidade["Descartados Semana"] / somaNotifSemana) * 100    
        ws['B25'] = "{:.2f}".format(porcSuspeitasSemana) + "%"
        ws['C25'] = "{:.2f}".format(porcPositivasSemana) + "%"
        ws['D25'] = "{:.2f}".format(porcNegativasSemana) + "%"    
        ws['E25'] = "{:.2f}".format(porcDescartadasSemana) + "%"
    ws['F24'] = somaNotifSemana    
    
    ws['A27'] = "Aplicações Vacivida (Total)"
    ws['B27'] = unidade["Vacivida Total"]
    ws['A28'] = "Aplicações Assessor (Total)"
    ws['B28'] = unidade["Saída de Vacina (Aplicação) Total"]
    
    ws['A30'] = "Aplicações Vacivida (Mes Passado)"
    ws['B30'] = unidade["Vacivida Mes Passado"]
    ws['A31'] = "Aplicações Assessor (Mes Passado)"
    ws['B31'] = unidade["Saída de Vacina (Aplicação) Mes Passado"]
    ws['A32'] = "Diferença (Assessor - Vacivida) (Mes Passado)"
    ws['B32'] = unidade["Aplicação Vacina Assessor - Vacivida (Mes Passado)"]
    
    ws['A34'] = "Aplicações Vacivida (Semana)"
    ws['B34'] = unidade["Vacivida Semana"]
    ws['A35'] = "Aplicações Assessor (Semana)"
    ws['B35'] = unidade["Saída de Vacina (Aplicação) Semana"]
    ws['A36'] = "Diferença (Assessor - Vacivida) (Semana)"
    ws['B36'] = unidade["Aplicação Vacina Assessor - Vacivida (Semana)"]
    
    ws['A38'] = "Testes rápidos dispensados (Total)"
    ws['B38'] = unidade["Testes Dispensados Total"]
    ws['A39'] = "Testes rápidos dispensados (Mes Passado)"
    ws['B39'] = unidade["Testes Dispensados Mes Passado"]
    ws['A40'] = "Testes rápidos dispensados (Semana)"
    ws['B40'] = unidade["Testes Dispensados Semana"]
    
    ws['G1'] = "TOTAL:"
    ws['G2'] = "01/01/2020"
    ws['H2'] = paramDataAtual.strftime("%d/%m/%Y")
    
    ws['G4'] = "MÊS PASSADO:"
    ws['G5'] = paramPrimeiroDiaMesPassado.strftime("%d/%m/%Y")
    ws['H5'] = paramUltimoDiaMesPassado.strftime("%d/%m/%Y")
    
    ws['G7'] = "SEMANA:"
    ws['G8'] = (paramDataAtual - timedelta(days=paramTamanhoSemana)).strftime("%d/%m/%Y")
    ws['H8'] = paramDataAtual.strftime("%d/%m/%Y")
    
    ws = autoSizeColumn(ws)    
    wb.save('Unidades/' + unidade['Apelido'] + " - " + paramDataAtras.strftime("%d.%m") + " a " + paramDataAtual.strftime("%d.%m") + ".xlsx")
    
consolidadoTestes(listaUnidades)